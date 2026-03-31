import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from io import BytesIO

from openpyxl import load_workbook


APP_TITLE = "CampusFlow Analytics"

REQUIRED_COLUMNS: dict[str, list[str]] = {
    "Resources": ["Room_ID", "Building", "Type", "Capacity"],
    "Schedule": ["Slot_ID", "Day", "Time", "Room_ID", "Course_ID"],
    "Utilization": ["Slot_ID", "Actual_Attendance", "Date"],
}

OPTIONAL_COLUMNS: dict[str, list[str]] = {
    "Schedule": ["Department"],
}

DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _require_columns(df: pd.DataFrame, sheet: str, required: list[str]) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Sheet '{sheet}' is missing required columns: {missing}")


@st.cache_data(show_spinner=False)
def load_workbook(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    xls = pd.ExcelFile(file_bytes)
    data: dict[str, pd.DataFrame] = {}
    for sheet, required_cols in REQUIRED_COLUMNS.items():
        if sheet not in xls.sheet_names:
            raise ValueError(f"Missing required sheet: '{sheet}'")
        df = pd.read_excel(xls, sheet_name=sheet)
        df = _normalize_columns(df)
        _require_columns(df, sheet, required_cols)

        keep_cols = list(required_cols)
        for extra in OPTIONAL_COLUMNS.get(sheet, []):
            if extra in df.columns:
                keep_cols.append(extra)

        data[sheet] = df[keep_cols].copy()
    return data


def update_utilization_attendance_xlsx(
    file_bytes: bytes,
    slot_id: str,
    new_attendance: int,
    *,
    update_mode: str = "latest_date",
) -> tuple[bytes, int]:
    """
    Update Actual_Attendance for a Slot_ID in the Utilization sheet.

    update_mode:
      - "latest_date": update only the row with the latest Date for that Slot_ID (preferred)
      - "all_rows": update all rows matching Slot_ID
    """
    if not slot_id or not str(slot_id).strip():
        raise ValueError("Slot_ID is required.")
    if new_attendance < 0:
        raise ValueError("Actual_Attendance must be >= 0.")

    slot_id = str(slot_id).strip()

    bio = BytesIO(file_bytes)
    wb = load_workbook(bio)
    if "Utilization" not in wb.sheetnames:
        raise ValueError("Workbook is missing 'Utilization' sheet.")

    ws = wb["Utilization"]
    header_row = None
    for r in range(1, min(10, ws.max_row) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if values and any(v is not None for v in values):
            header_row = r
            break

    if header_row is None:
        raise ValueError("'Utilization' sheet appears to be empty.")

    headers = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(row=header_row, column=c).value
        if name is None:
            continue
        headers[str(name).strip()] = c

    for required in REQUIRED_COLUMNS["Utilization"]:
        if required not in headers:
            raise ValueError(f"'Utilization' sheet is missing required column '{required}'.")

    col_slot = headers["Slot_ID"]
    col_att = headers["Actual_Attendance"]
    col_date = headers["Date"]

    matches: list[tuple[int, object]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_slot).value
        if v is None:
            continue
        if str(v).strip() == slot_id:
            matches.append((r, ws.cell(row=r, column=col_date).value))

    if not matches:
        raise ValueError(f"Slot_ID '{slot_id}' not found in Utilization.")

    rows_to_update: list[int]
    if update_mode == "all_rows":
        rows_to_update = [r for r, _ in matches]
    elif update_mode == "latest_date":
        # If date parsing fails, fall back to last matched row.
        def _as_ts(x: object) -> pd.Timestamp:
            return pd.to_datetime(x, errors="coerce")

        parsed = [(r, _as_ts(d)) for r, d in matches]
        parsed_valid = [(r, d) for r, d in parsed if pd.notna(d)]
        if parsed_valid:
            best_row = max(parsed_valid, key=lambda t: t[1])[0]
        else:
            best_row = matches[-1][0]
        rows_to_update = [best_row]
    else:
        raise ValueError("update_mode must be 'latest_date' or 'all_rows'.")

    for r in rows_to_update:
        ws.cell(row=r, column=col_att).value = int(new_attendance)

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), len(rows_to_update)


def underutilized_slots(fact: pd.DataFrame, threshold: float = 0.30) -> pd.DataFrame:
    df = fact.dropna(subset=["Capacity", "Actual_Attendance"]).copy()
    df = df[df["Capacity"] > 0].copy()
    df["Underutilized"] = (df["Actual_Attendance"] < (threshold * df["Capacity"])).fillna(False)
    return df[df["Underutilized"]].copy()


def recommend_moves(
    under: pd.DataFrame, resources: pd.DataFrame, sqft_per_seat: float = 20.0
) -> tuple[pd.DataFrame, float]:
    """
    Suggest moving underutilized classes to smaller rooms that are available at the same Date/Day/Time.
    Returns (recommendations_df, potential_space_recovery_sqft).
    """
    if under.empty:
        return under.copy(), 0.0

    r = resources.copy()
    r["Room_ID"] = r["Room_ID"].astype(str).str.strip()
    r["Capacity"] = pd.to_numeric(r["Capacity"], errors="coerce")

    # Room availability at a specific Date/Day/Time
    occupied = (
        under[["Date", "Day", "Time", "Room_ID"]]
        .dropna(subset=["Date", "Day", "Time", "Room_ID"])
        .assign(Room_ID=lambda x: x["Room_ID"].astype(str).str.strip())
    )
    # NOTE: under is already a subset of fact; but we need occupancy from *all* sessions.
    # We'll reconstruct occupancy by taking distinct room use from the full fact via columns present in under.
    # (Caller should pass under derived from full fact; we use the parent frame below in main.)

    recs: list[dict] = []
    recovery_sqft = 0.0

    # We'll precompute candidate lists per type for speed.
    by_type: dict[str, pd.DataFrame] = {
        t: grp.sort_values("Capacity").copy() for t, grp in r.dropna(subset=["Capacity"]).groupby("Type")
    }
    all_rooms_sorted = r.dropna(subset=["Capacity"]).sort_values("Capacity").copy()

    for _, row in under.iterrows():
        cur_room = str(row.get("Room_ID", "")).strip()
        cur_type = row.get("Type")
        cur_cap = row.get("Capacity")
        att = row.get("Actual_Attendance")

        if not np.isfinite(cur_cap) or not np.isfinite(att) or cur_cap <= 0:
            continue

        # Determine which rooms are occupied in the same Date/Day/Time in the *full* dataset
        occ_rooms = row.get("_occupied_rooms_set")
        if not isinstance(occ_rooms, set):
            occ_rooms = set()

        # Type preference: same type first, else any type.
        candidates = by_type.get(cur_type, all_rooms_sorted)
        # Room must:
        # - be different
        # - be free at that Date/Day/Time
        # - have capacity >= attendance
        # - be smaller than current room capacity
        cand = candidates[
            (candidates["Room_ID"] != cur_room)
            & (~candidates["Room_ID"].isin(list(occ_rooms)))
            & (candidates["Capacity"] >= att)
            & (candidates["Capacity"] < cur_cap)
        ].copy()

        if cand.empty and candidates is not all_rooms_sorted:
            cand = all_rooms_sorted[
                (all_rooms_sorted["Room_ID"] != cur_room)
                & (~all_rooms_sorted["Room_ID"].isin(list(occ_rooms)))
                & (all_rooms_sorted["Capacity"] >= att)
                & (all_rooms_sorted["Capacity"] < cur_cap)
            ].copy()

        if cand.empty:
            continue

        # Choose the smallest room that still fits attendance.
        best = cand.sort_values("Capacity").iloc[0]
        new_room = str(best["Room_ID"])
        new_cap = float(best["Capacity"])

        recovered_seats = float(cur_cap - new_cap)
        recovered_sqft = max(0.0, recovered_seats * float(sqft_per_seat))
        recovery_sqft += recovered_sqft

        recs.append(
            {
                "Date": row.get("Date"),
                "Day": row.get("Day"),
                "Time": row.get("Time"),
                "Course_ID": row.get("Course_ID"),
                "Department": row.get("Department", np.nan),
                "Current_Room_ID": cur_room,
                "Current_Capacity": cur_cap,
                "Actual_Attendance": att,
                "Recommended_Room_ID": new_room,
                "Recommended_Capacity": new_cap,
                "Recovered_SqFt": recovered_sqft,
            }
        )

    rec_df = pd.DataFrame(recs)
    if not rec_df.empty:
        rec_df = rec_df.sort_values(["Recovered_SqFt", "Current_Capacity"], ascending=[False, False])
    return rec_df, float(recovery_sqft)


def build_fact_table(resources: pd.DataFrame, schedule: pd.DataFrame, utilization: pd.DataFrame) -> pd.DataFrame:
    r = resources.copy()
    s = schedule.copy()
    u = utilization.copy()

    r["Room_ID"] = r["Room_ID"].astype(str).str.strip()
    s["Room_ID"] = s["Room_ID"].astype(str).str.strip()
    s["Slot_ID"] = s["Slot_ID"].astype(str).str.strip()
    u["Slot_ID"] = u["Slot_ID"].astype(str).str.strip()

    r["Capacity"] = pd.to_numeric(r["Capacity"], errors="coerce")
    u["Actual_Attendance"] = pd.to_numeric(u["Actual_Attendance"], errors="coerce")
    u["Date"] = pd.to_datetime(u["Date"], errors="coerce")

    s["Day"] = s["Day"].astype(str).str.strip()
    s["Time"] = s["Time"].astype(str).str.strip()

    fact = s.merge(u, on="Slot_ID", how="left", validate="many_to_many")
    fact = fact.merge(r, on="Room_ID", how="left", validate="many_to_one")

    # Core metrics
    fact["OverCapacity"] = (fact["Actual_Attendance"] > fact["Capacity"]).fillna(False)

    util_ratio = fact["Actual_Attendance"] / fact["Capacity"]
    fact["Utilization_Ratio"] = util_ratio.replace([np.inf, -np.inf], np.nan)

    # Revolutionary feature: Efficiency Score per room+slot record
    # - Primary signal: attendance vs physical capacity
    # - Display capped at 100 (over-capacity is handled via alert)
    fact["Efficiency_Score"] = (100.0 * fact["Utilization_Ratio"]).clip(lower=0, upper=100)

    # Ghost rooms: scheduled but largely empty
    fact["GhostRoom"] = (fact["Utilization_Ratio"] < 0.20).fillna(False)

    # Waste index: empty seats, scaled by how "wasteful" the booking is
    fact["Empty_Seats"] = (fact["Capacity"] - fact["Actual_Attendance"]).clip(lower=0)
    fact["Waste_Index"] = (fact["Empty_Seats"] * (1 - fact["Utilization_Ratio"].fillna(0))).fillna(0)

    return fact


def compute_global_utilization_pct(fact: pd.DataFrame) -> float | None:
    df = fact.dropna(subset=["Capacity", "Actual_Attendance"]).copy()
    if df.empty:
        return None
    cap = df["Capacity"].sum()
    if not np.isfinite(cap) or cap <= 0:
        return None
    return float(100.0 * df["Actual_Attendance"].sum() / cap)


def pick_five_times(fact: pd.DataFrame) -> list[str]:
    times = (
        fact["Time"]
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", np.nan)
        .dropna()
        .value_counts()
        .index.tolist()
    )
    if len(times) >= 5:
        return times[:5]
    # Pad deterministically so the heatmap is always 5 rows
    return (times + [f"(Empty {i})" for i in range(1, 6)])[:5]


def heatmap_data(fact: pd.DataFrame, times_5: list[str]) -> pd.DataFrame:
    df = fact.copy()
    df["Day"] = pd.Categorical(df["Day"], categories=DAY_ORDER, ordered=True)

    df = df[df["Day"].isin(DAY_ORDER)].copy()
    df = df[df["Time"].isin(times_5)].copy()

    # Heat metric blends overcrowding and dead zones:
    # - High positive: overcrowding (ratio > 1)
    # - Negative: underuse (ratio near 0)
    # We use (ratio - 0.5) to center around "acceptable" ~50% use.
    df["Heat_Signal"] = (df["Utilization_Ratio"] - 0.5).clip(lower=-0.5, upper=1.5)

    pivot = (
        df.pivot_table(
            index="Time",
            columns="Day",
            values="Heat_Signal",
            aggfunc="mean",
        )
        .reindex(index=times_5, columns=DAY_ORDER)
        .fillna(0)
    )
    pivot.index.name = "Time"
    return pivot


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")

    st.title(APP_TITLE)
    st.caption("Excel-backed campus resource efficiency analytics with per-room, per-slot **Efficiency Score**.")

    with st.sidebar:
        st.subheader("Excel Database")
        uploaded = st.file_uploader("Upload `.xlsx`", type=["xlsx"])
        if uploaded is not None:
            # Persist the uploaded workbook bytes so we can mutate the "Excel backend" in-session.
            st.session_state["excel_name"] = uploaded.name
            st.session_state["excel_bytes"] = uploaded.getvalue()

        excel_bytes = st.session_state.get("excel_bytes", None)
        excel_name = st.session_state.get("excel_name", "Campus_Data.xlsx")

        if excel_bytes is not None:
            st.download_button(
                "Download updated Excel",
                data=excel_bytes,
                file_name=excel_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()
        st.subheader("Efficiency Settings")
        ghost_threshold = st.slider("Ghost room threshold (utilization < x)", 0.05, 0.50, 0.20, 0.01)
        underutil_threshold = st.slider("Underutilized threshold (utilization < x)", 0.05, 0.60, 0.30, 0.01)
        sqft_per_seat = st.slider("Sq ft per seat (assumption)", 10, 35, 20, 1)

        st.divider()
        st.subheader("Manual Attendance Update")
        if excel_bytes is None:
            st.caption("Upload an Excel file to enable manual updates.")
        else:
            # Read Slot_IDs from current in-session backend.
            try:
                cur_wb = load_workbook(excel_bytes)
                util_df = cur_wb["Utilization"].copy()
                slot_options = (
                    util_df["Slot_ID"]
                    .astype(str)
                    .str.strip()
                    .replace("", np.nan)
                    .dropna()
                    .unique()
                    .tolist()
                )
                slot_options = sorted(slot_options)
            except Exception:
                slot_options = []

            slot_choice = st.selectbox(
                "Select Slot_ID",
                options=slot_options,
                index=0 if slot_options else None,
            )
            new_att = st.number_input("Actual_Attendance", min_value=0, step=1, value=30)
            update_mode = st.radio(
                "Update scope",
                options=["Update latest date only", "Update all matching rows"],
                index=0,
            )

            if st.button("Update Excel Backend", use_container_width=True, type="primary"):
                mode = "latest_date" if update_mode.startswith("Update latest") else "all_rows"
                try:
                    updated_bytes, n = update_utilization_attendance_xlsx(
                        excel_bytes,
                        slot_id=str(slot_choice),
                        new_attendance=int(new_att),
                        update_mode=mode,
                    )
                    st.session_state["excel_bytes"] = updated_bytes
                    st.success(f"Updated {n} row(s) for Slot_ID {slot_choice}. Refreshing analytics…")
                    st.rerun()
                except Exception as e:
                    st.error(f"Update failed: {e}")

    if not uploaded and st.session_state.get("excel_bytes", None) is None:
        st.info("Upload your multi-sheet Excel file to begin.")
        return

    try:
        wb = load_workbook(st.session_state["excel_bytes"])
    except Exception as e:
        st.error(f"Could not read workbook: {e}")
        return

    fact = build_fact_table(wb["Resources"], wb["Schedule"], wb["Utilization"])
    fact["GhostRoom"] = (fact["Utilization_Ratio"] < float(ghost_threshold)).fillna(False)

    # Data validation: over-capacity alerts
    over = fact[fact["OverCapacity"]].copy()
    if not over.empty:
        st.warning(f"Over-capacity detected: {len(over)} record(s) where Attendance > Capacity.")
        st.dataframe(
            over[
                [
                    "Date",
                    "Day",
                    "Time",
                    "Building",
                    "Room_ID",
                    "Type",
                    "Course_ID",
                    "Capacity",
                    "Actual_Attendance",
                ]
            ].sort_values(["Date", "Day", "Time"], na_position="last"),
            use_container_width=True,
        )

    # Top section: global metric
    util_pct = compute_global_utilization_pct(fact)
    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        st.metric("Global Campus Utilization %", "—" if util_pct is None else f"{util_pct:.1f}%")
    with c2:
        ghost_count = int(fact["GhostRoom"].fillna(False).sum())
        st.metric("Ghost Room Records", f"{ghost_count}")
    with c3:
        st.caption(
            "Global utilization is computed from records where both `Capacity` and `Actual_Attendance` exist. "
            "Efficiency Score is computed as 100 × (Attendance / Capacity) (capped at 100 for display)."
        )

    st.divider()

    left, right = st.columns([1.15, 0.85])

    with left:
        st.subheader("Heatmap (Mon–Fri × 5 Times)")
        default_times = pick_five_times(fact)
        all_times = (
            fact["Time"]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", np.nan)
            .dropna()
            .unique()
            .tolist()
        )
        all_times_sorted = sorted(all_times)

        # Keep the grid strictly 5x5, but allow choosing which 5 "Time" values drive it.
        selected_times = st.multiselect(
            "Pick exactly 5 time labels for the grid",
            options=all_times_sorted,
            default=[t for t in default_times if t in all_times_sorted],
            max_selections=5,
        )
        if len(selected_times) != 5:
            selected_times = (selected_times + default_times)[:5]

        pivot = heatmap_data(fact, selected_times)

        fig = px.imshow(
            pivot,
            color_continuous_scale="RdBu",
            origin="lower",
            aspect="auto",
            labels={"color": "Overcrowding ↔ Dead Zone"},
        )
        fig.update_layout(
            margin=dict(l=10, r=10, t=30, b=10),
            coloraxis_colorbar=dict(
                tickvals=[-0.5, 0.0, 0.5, 1.0, 1.5],
                ticktext=["Dead", "Low", "OK", "High", "Over"],
            ),
        )
        st.plotly_chart(fig, use_container_width=True)

    with right:
        st.subheader("Waste Finder (Top 5)")
        df = fact.dropna(subset=["Capacity", "Actual_Attendance"]).copy()
        if df.empty:
            st.info("Not enough utilization data to compute waste.")
        else:
            waste = (
                df.assign(
                    Efficiency_Score=lambda x: (100.0 * (x["Actual_Attendance"] / x["Capacity"]))
                    .replace([np.inf, -np.inf], np.nan)
                    .clip(0, 100)
                )
                .sort_values(["Waste_Index", "Capacity"], ascending=[False, False])
                .head(5)
            )
            st.dataframe(
                waste[
                    [
                        "Date",
                        "Day",
                        "Time",
                        "Building",
                        "Room_ID",
                        "Type",
                        "Course_ID",
                        "Capacity",
                        "Actual_Attendance",
                        "Efficiency_Score",
                        "GhostRoom",
                    ]
                ],
                use_container_width=True,
            )

    st.divider()
    st.subheader("Unified Data (Joined)")
    st.dataframe(
        fact[
            [
                "Date",
                "Slot_ID",
                "Day",
                "Time",
                "Building",
                "Room_ID",
                "Type",
                "Course_ID",
                "Capacity",
                "Actual_Attendance",
                "Utilization_Ratio",
                "Efficiency_Score",
                "GhostRoom",
                "OverCapacity",
            ]
        ].sort_values(["Date", "Day", "Time"], na_position="last"),
        use_container_width=True,
        height=340,
    )

    st.divider()
    st.subheader("Strategic Optimization Suggestions")

    under = underutilized_slots(fact, threshold=float(underutil_threshold))
    st.caption(
        "Underutilized slots are sessions where Actual Attendance is below the selected share of physical capacity. "
        "Recommendations move them to smaller rooms that are free at the same Date/Day/Time."
    )

    if under.empty:
        st.info("No underutilized slots found (given the current threshold and available data).")
    else:
        # Build occupancy map from the full fact table (Date/Day/Time -> rooms occupied).
        occ = (
            fact.dropna(subset=["Date", "Day", "Time", "Room_ID"])[["Date", "Day", "Time", "Room_ID"]]
            .assign(Room_ID=lambda x: x["Room_ID"].astype(str).str.strip())
            .drop_duplicates()
        )
        occ_map = (
            occ.groupby(["Date", "Day", "Time"])["Room_ID"]
            .apply(lambda s: set(s.tolist()))
            .to_dict()
        )

        under = under.copy()
        under["_occupied_rooms_set"] = under.apply(
            lambda r: occ_map.get((r.get("Date"), r.get("Day"), r.get("Time")), set()),
            axis=1,
        )

        rec_df, recovery_sqft = recommend_moves(
            under=under,
            resources=wb["Resources"],
            sqft_per_seat=float(sqft_per_seat),
        )

        k1, k2, k3 = st.columns([1, 1, 2])
        with k1:
            st.metric("Underutilized Slot Records", f"{len(under):,}")
        with k2:
            st.metric("Move Recommendations", f"{len(rec_df):,}")
        with k3:
            st.metric("Potential Space Recovery (sq ft)", f"{recovery_sqft:,.0f}")

        if rec_df.empty:
            st.info("No feasible moves found (no smaller available rooms that fit attendance at the same times).")
        else:
            st.dataframe(
                rec_df[
                    [
                        "Date",
                        "Day",
                        "Time",
                        "Department",
                        "Course_ID",
                        "Current_Room_ID",
                        "Current_Capacity",
                        "Actual_Attendance",
                        "Recommended_Room_ID",
                        "Recommended_Capacity",
                        "Recovered_SqFt",
                    ]
                ].head(25),
                use_container_width=True,
            )

    st.divider()
    st.subheader("Departmental Fairness")

    if "Department" not in fact.columns:
        st.info("No `Department` column found in the uploaded Schedule sheet, so fairness charts are unavailable.")
    else:
        dept_df = fact.dropna(subset=["Department", "Capacity", "Actual_Attendance"]).copy()
        if dept_df.empty:
            st.info("Not enough departmental data to compute fairness metrics.")
        else:
            fairness = (
                dept_df.groupby("Department", dropna=True)
                .agg(
                    Booked_Capacity=("Capacity", "sum"),
                    Student_Footprint=("Actual_Attendance", "sum"),
                    Sessions=("Slot_ID", "count"),
                )
                .reset_index()
            )
            fairness["Utilization_%"] = np.where(
                fairness["Booked_Capacity"] > 0,
                100.0 * fairness["Student_Footprint"] / fairness["Booked_Capacity"],
                np.nan,
            )

            fig = px.bar(
                fairness.sort_values("Booked_Capacity", ascending=False),
                x="Department",
                y=["Booked_Capacity", "Student_Footprint"],
                barmode="group",
                title="Booked Capacity vs Student Footprint",
                labels={"value": "Total (seat-count)", "variable": "Metric"},
            )
            st.plotly_chart(fig, use_container_width=True)

            fig2 = px.scatter(
                fairness,
                x="Booked_Capacity",
                y="Student_Footprint",
                size="Sessions",
                color="Department",
                hover_data=["Utilization_%"],
                title="Resource Use vs Student Footprint (by department)",
                labels={"Booked_Capacity": "Booked capacity (sum)", "Student_Footprint": "Attendance (sum)"},
            )
            st.plotly_chart(fig2, use_container_width=True)

            st.dataframe(
                fairness.sort_values("Booked_Capacity", ascending=False),
                use_container_width=True,
            )


if __name__ == "__main__":
    main()

